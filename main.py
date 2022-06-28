from matplotlib.pyplot import contour
from numpy import average
import openpyxl as xl
from openpyxl.styles import Side, Border, PatternFill, Color, Font, Alignment
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
import pandas as pd
import tkinter as tk
from tkinter import filedialog as fd
from tkinter import messagebox as msg

class MyFrame(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)

        self.master = master
        self.master.title("성호를 위한 작은 선물")
        self.pack(fill=tk.BOTH, expand=True)

        #(왼쪽) 결과 미리보기창
        Frame1 = tk.Frame(self)
        Frame1.pack(side=tk.LEFT)

        Frame3 = tk.Frame(Frame1)
        Frame3.pack(side=tk.TOP)

        ColorLbl1 = tk.Label(Frame3, text="   학생점수:")
        ColorLbl1.pack(side=tk.LEFT, padx=5, pady=5)
        ColorCheck1 = tk.Checkbutton(Frame3, text="파랑")
        ColorCheck2 = tk.Checkbutton(Frame3, text="빨강")
        ColorCheck3 = tk.Checkbutton(Frame3, text="노랑")
        ColorCheck1.pack(side=tk.LEFT, padx=5, pady=5)
        ColorCheck2.pack(side=tk.LEFT, padx=5, pady=5)
        ColorCheck3.pack(side=tk.LEFT, padx=5, pady=5)

        ColorLbl2 = tk.Label(Frame3, text="   테스트 평균:")
        ColorLbl2.pack(side=tk.LEFT, padx=5, pady=5)
        ColorCheck4 = tk.Checkbutton(Frame3, text="파랑")
        ColorCheck5 = tk.Checkbutton(Frame3, text="빨강")
        ColorCheck6 = tk.Checkbutton(Frame3, text="노랑")
        ColorCheck4.pack(side=tk.LEFT, padx=5, pady=5)
        ColorCheck5.pack(side=tk.LEFT, padx=5, pady=5)
        ColorCheck6.pack(side=tk.LEFT, padx=5, pady=5)

        Canvas = tk.Frame(Frame1, relief='solid', bd=2, width=470, height=470)
        Canvas.pack(side=tk.BOTTOM, padx=10, pady=10)

        #(오른쪽) 버튼 모음
        Frame2 = tk.Frame(self)
        Frame2.pack(side=tk.RIGHT)

        imptBtn = tk.Button(Frame2, text="불러오기", width=20, height=10, command=self.imptFile)
        imptBtn.pack(side=tk.TOP, padx=10, pady=20)

        exptBtn = tk.Button(Frame2, text="내보내기", width=20, height=10, command=self.exptFile)
        exptBtn.pack(side=tk.TOP, padx=10, pady=20)

    def imptFile(self):
        f = fd.askopenfilename(title="불러오기", filetypes = (("Excel files", ".xlsx .xls"),))
        self.readFile(f)

    def exptFile(self):
        name = fd.asksaveasfilename(title="내보내기", filetypes=(("Excel files", ".xlsx .xls"),))
        try:
            self.WB.save(name+".xlsx")
            self.WB.close()
        except:
            msg.WARNING("엑셀파일을 먼저 불러오세요")

    def readFile(self, f):
        wb = xl.load_workbook(f)
        ws =wb[wb.sheetnames[0]]
        self.i = 2
        self.j = 1

        self.WB = xl.Workbook()
        self.WS = self.WB.active
        R = 2
        C = 2
        while self.i < 100:
            goOn = self.findTable(ws)

            if goOn:
                column = self.df.shape[1]
                row = self.df.shape[0]
                self.avgList = self.getAverage()

                for i in range(row):
                    self.createTable(i, R, C, row, column)
                    R += column+10

            else:
                break

        #대충 왼쪽 창에 결과물 보여주는 코드 추가필요
        wb.close()

    def findTable(self, sheet):
        while sheet.cell(row=self.i-self.j,column=self.j).value != "이름":
            self.j -= 1
            if self.j == 0:
                self.i += 1
                self.j = self.i-1
            if self.i > 100:
                return False
        if self.j == 0:
            self.i += 1
            self.j = self.i-1

        rowCnt = 0
        columnCnt = 0

        self.df = pd.DataFrame()
        personalData = []
        self.dateList = []

        while sheet.cell(row=self.i-self.j+rowCnt, column=self.j+columnCnt).value:
            personalData.append(sheet.cell(row=self.i-self.j+rowCnt,column=self.j+columnCnt).value)
            rowCnt += 1
            if not sheet.cell(row=self.i-self.j+rowCnt, column=self.j+columnCnt).value:
                self.df[str(personalData[0])] = personalData[1:rowCnt]
                columnCnt += 1
                rowCnt = 0
                self.dateList.append(personalData[0])
                personalData = []
        self.dateList.remove(self.dateList[0])
        self.j -= 1
        
        return True

    def getAverage(self):
        averageList = []

        for j in range(len(self.dateList)):
            sum = 0
            cnt = 0
            for i in range(len(self.df)):
                if self.df.iloc[i][str(self.dateList[j])]=='미응시':
                    continue
                else:
                    sum += self.df.iloc[i][str(self.dateList[j])]
                    cnt += 1
            avg = sum/cnt
            avg = "%0.1f" %avg
            avg = float(avg)
            averageList.append(avg)
        return averageList

    def createTable(self, studentNumber, X, Y, row, column):
        self.WS.merge_cells(start_row=X,start_column=Y,end_row=X,end_column=Y+8)
        self.WS.cell(row=X, column=Y).value = "수학원정대 토요 클리닉 월별 점수"
        self.WS.cell(row=X, column=Y).alignment = Alignment(horizontal='center',
                                                            vertical='center')
        self.WS.cell(row=X, column=Y).font = Font(bold=True,
                                                    color='FFFFFF')
        self.WS.cell(row=X, column=Y).fill = PatternFill(fill_type='solid',
                                                            fgColor=Color('123456'))

        self.WS.merge_cells(start_row=X+1,start_column=Y,end_row=X+1,end_column=Y+8)
        self.WS.cell(row=X+1, column=Y).value = "지역 최상위로 키웁니다. 수학원정대 보습학원"
        self.WS.cell(row=X+1, column=Y).font = Font(bold=True,
                                                    size=8)
        self.WS.cell(row=X+1, column=Y).alignment = Alignment(horizontal='right',
                                                                vertical='bottom')

        self.WS.merge_cells(start_row=X+2,start_column=Y,end_row=X+3,end_column=Y)
        self.WS.cell(row=X+2, column=Y).value = "이름"
        self.WS.cell(row=X+2, column=Y).font = Font(bold=True)
        self.WS.cell(row=X+2, column=Y).alignment = Alignment(horizontal='center',
                                                            vertical='center')
        self.WS.cell(row=X+2, column=Y).fill = PatternFill(fill_type='solid',
                                                            fgColor=Color('00FFFF'))

        self.WS.merge_cells(start_row=X+2,start_column=Y+1,end_row=X+3,end_column=Y+2)
        self.WS.cell(row=X+2, column=Y+1).value = self.df.iloc[studentNumber]['이름']
        self.WS.cell(row=X+2, column=Y+1).font = Font(bold=True)
        self.WS.cell(row=X+2, column=Y+1).alignment = Alignment(horizontal='center',
                                                                vertical='center')

        self.WS.cell(row=X+2, column=Y+3).value="학교"
        self.WS.cell(row=X+2, column=Y+5).value="반명"
        self.WS.cell(row=X+2, column=Y+7).value="Test 과정"
        self.WS.cell(row=X+3, column=Y+3).value="학년"
        self.WS.cell(row=X+3, column=Y+5).value="진행기간"
        self.WS.cell(row=X+3, column=Y+7).value="담임"
        self.WS.cell(row=X+2, column=Y+3).font=Font(bold=True)
        self.WS.cell(row=X+2, column=Y+5).font=Font(bold=True)
        self.WS.cell(row=X+2, column=Y+7).font=Font(bold=True)
        self.WS.cell(row=X+3, column=Y+3).font=Font(bold=True)
        self.WS.cell(row=X+3, column=Y+5).font=Font(bold=True)
        self.WS.cell(row=X+3, column=Y+7).font=Font(bold=True)
        self.WS.cell(row=X+2, column=Y+3).alignment=Alignment(horizontal='left',
                                                            vertical='center')
        self.WS.cell(row=X+2, column=Y+5).alignment=Alignment(horizontal='left',
                                                            vertical='center')
        self.WS.cell(row=X+2, column=Y+7).alignment=Alignment(horizontal='left',
                                                            vertical='center')
        self.WS.cell(row=X+3, column=Y+3).alignment=Alignment(horizontal='left',
                                                            vertical='center')
        self.WS.cell(row=X+3, column=Y+5).alignment=Alignment(horizontal='left',
                                                            vertical='center')
        self.WS.cell(row=X+3, column=Y+7).alignment=Alignment(horizontal='left',
                                                            vertical='center')
        self.WS.cell(row=X+2, column=Y+3).fill= PatternFill(fill_type='solid',
                                                            fgColor=Color('00FFFF'))
        self.WS.cell(row=X+2, column=Y+5).fill= PatternFill(fill_type='solid',
                                                            fgColor=Color('00FFFF'))
        self.WS.cell(row=X+2, column=Y+7).fill= PatternFill(fill_type='solid',
                                                            fgColor=Color('00FFFF'))
        self.WS.cell(row=X+3, column=Y+3).fill= PatternFill(fill_type='solid',
                                                            fgColor=Color('00FFFF'))
        self.WS.cell(row=X+3, column=Y+5).fill= PatternFill(fill_type='solid',
                                                            fgColor=Color('00FFFF'))
        self.WS.cell(row=X+3, column=Y+7).fill= PatternFill(fill_type='solid',
                                                            fgColor=Color('00FFFF'))

        self.WS.merge_cells(start_row=X+4,start_column=Y,end_row=X+4,end_column=Y+8)

        self.WS.cell(row=X+5, column=Y).value="날짜"
        self.WS.cell(row=X+5, column=Y+1).value="단원"
        self.WS.cell(row=X+5, column=Y+2).value="점수"
        self.WS.cell(row=X+5, column=Y+3).value="반 평균"
        self.WS.cell(row=X+5, column=Y).font=Font(bold=True)
        self.WS.cell(row=X+5, column=Y+1).font=Font(bold=True)
        self.WS.cell(row=X+5, column=Y+2).font=Font(bold=True)
        self.WS.cell(row=X+5, column=Y+3).font=Font(bold=True)
        self.WS.cell(row=X+5, column=Y).alignment=Alignment(horizontal='center',
                                                                vertical='center')
        self.WS.cell(row=X+5, column=Y+1).alignment=Alignment(horizontal='center',
                                                                vertical='center')
        self.WS.cell(row=X+5, column=Y+2).alignment=Alignment(horizontal='center',
                                                                vertical='center')
        self.WS.cell(row=X+5, column=Y+3).alignment=Alignment(horizontal='center',
                                                                vertical='center')

        for i in range(len(self.dateList)):
            scoreList = []
            scoreList.append(self.df.iloc[studentNumber][str(self.dateList[i])])
            minScore = min(scoreList)

            self.WS.cell(row=X+6+i, column=Y).value = self.dateList[i].strftime("%m/%d".encode('unicode-escape').decode()).encode().decode('unicode-escape')
            self.WS.cell(row=X+6+i, column=Y+2).value = self.df.iloc[studentNumber][str(self.dateList[i])]
            self.WS.cell(row=X+6+i, column=Y+3).value = self.avgList[i]
            self.WS.cell(row=X+6+i, column=Y).alignment = Alignment(horizontal='center',
                                                                    vertical='center')
            self.WS.cell(row=X+6+i, column=Y+2).alignment = Alignment(horizontal='center',
                                                                        vertical='center')
            self.WS.cell(row=X+6+i, column=Y+3).alignment = Alignment(horizontal='center',
                                                                        vertical='center')
            self.WS.cell(row=X+6+i, column=Y).font = Font(bold=True)
            self.WS.cell(row=X+6+i, column=Y+2).font = Font(bold=True)
            self.WS.cell(row=X+6+i, column=Y+3).font = Font(bold=True)

        self.WS.merge_cells(start_row=X+5,start_column=Y+5,end_row=X+5+column,end_column=Y+8)
        BarPos = self.WS.cell(row=X+5, column=Y+5).coordinate
        BarCats = Reference(self.WS, min_row=X+6, min_col=Y, max_row=X+6+len(self.dateList), max_col=Y)
        ScoreData = Reference(self.WS, min_row=X+6, min_col=Y+2, max_row=X+6+len(self.dateList), max_col=Y+2)
        ScoreChart = BarChart()
        ScoreChart.add_data(ScoreData)
        ScoreChart.set_categories(BarCats)
        ScoreChart.y_axis.scaling.max = 100
        ScoreChart.width = 4/0.55
        ScoreChart.height = (len(self.dateList)+2)/1.8


        AverageChart = LineChart()
        AverageData = Reference(self.WS, min_row=X+6, min_col=Y+3, max_row=X+6+len(self.dateList), max_col=Y+3)
        AverageChart.add_data(AverageData)
        ScoreChart += AverageChart
        ScoreChart.y_axis.scaling.min = max(min(minScore, min(self.avgList))-10, 0)
        ScoreChart.legend = None

        self.WS.add_chart(ScoreChart, BarPos)
        


        for i in range(9):
            for j in range(6+column):
                self.WS.cell(row=X+j, column=Y+i).border = Border(left=Side(border_style='medium', color='000000'),
                                                                right=Side(border_style='medium', color='000000'),
                                                                top=Side(border_style='medium', color='000000'),
                                                                bottom=Side(border_style='medium', color='000000'))


def main():
        root = tk.Tk()
        root.geometry("650x550+500+100")
        Frame = MyFrame(root)

        root.mainloop()



if __name__ == '__main__':
    main()